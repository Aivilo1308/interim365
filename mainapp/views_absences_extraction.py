# -*- coding: utf-8 -*-
"""
Vues pour l'extraction des absences utilisateurs avec jours fériés

Fichier à placer dans: mainapp/views_absences_extraction.py

Fonctionnalités:
- Liste des absences avec indication des jours fériés dans la période
- Filtrage par département, site, matricule, période
- Export PDF, XLSX, CSV
- Logging avancé pour audit et détection d'anomalies
"""

from datetime import date, datetime, timedelta
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
import io
import csv
import time
import logging

# ================================================================
# CONFIGURATION LOGGING
# ================================================================

logger = logging.getLogger('interim')
action_logger = logging.getLogger('interim.actions')
anomaly_logger = logging.getLogger('interim.anomalies')
perf_logger = logging.getLogger('interim.performance')


def log_action(category, action, message, request=None, **kwargs):
    """Log une action utilisateur avec contexte"""
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    user_info = "anonymous"
    ip_addr = "-"
    
    if request and hasattr(request, 'user') and request.user.is_authenticated:
        user_info = request.user.username
        ip_addr = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '-'))
        if ',' in ip_addr:
            ip_addr = ip_addr.split(',')[0].strip()
    
    extra_info = ' '.join([f"[{k}:{v}]" for k, v in kwargs.items() if v is not None])
    log_msg = f"[{timestamp}] [{category}] [{action}] [User:{user_info}] [IP:{ip_addr}] {extra_info} {message}"
    
    action_logger.info(log_msg)
    logger.info(log_msg)


def log_anomalie(category, message, severite='WARNING', request=None, **kwargs):
    """Log une anomalie détectée"""
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    user_info = "anonymous"
    
    if request and hasattr(request, 'user') and request.user.is_authenticated:
        user_info = request.user.username
    
    extra_info = ' '.join([f"[{k}:{v}]" for k, v in kwargs.items() if v is not None])
    log_msg = f"[{timestamp}] [ANOMALIE] [{category}] [{severite}] [User:{user_info}] {extra_info} {message}"
    
    if severite == 'ERROR':
        anomaly_logger.error(f"❌ {log_msg}")
        logger.error(f"❌ ANOMALIE: {log_msg}")
    elif severite == 'CRITICAL':
        anomaly_logger.critical(f"🔥 {log_msg}")
        logger.critical(f"🔥 ANOMALIE CRITIQUE: {log_msg}")
    else:
        anomaly_logger.warning(f"⚠️ {log_msg}")
        logger.warning(f"⚠️ ANOMALIE: {log_msg}")


def log_resume(operation, stats, duree_ms=None):
    """Log un résumé d'opération avec statistiques"""
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    lines = [
        "",
        "=" * 60,
        f"📊 RÉSUMÉ: {operation}",
        "=" * 60,
        f"⏰ Date/Heure: {timestamp}",
    ]
    
    if duree_ms is not None:
        if duree_ms >= 60000:
            duree_str = f"{duree_ms/60000:.1f} min"
        elif duree_ms >= 1000:
            duree_str = f"{duree_ms/1000:.1f} sec"
        else:
            duree_str = f"{duree_ms:.0f} ms"
        lines.append(f"⏱️ Durée: {duree_str}")
    
    lines.append("📈 Statistiques:")
    for key, value in stats.items():
        icon = '✅' if 'succes' in key.lower() or 'ok' in key.lower() else \
               '❌' if 'erreur' in key.lower() or 'echec' in key.lower() else \
               '⚠️' if 'warning' in key.lower() or 'anomalie' in key.lower() else '•'
        lines.append(f"   {icon} {key}: {value}")
    
    # Statut global
    erreurs = stats.get('erreurs', 0) + stats.get('echecs', 0)
    if erreurs == 0:
        lines.append("✅ Statut: SUCCÈS")
    elif erreurs > 5:
        lines.append("❌ Statut: ÉCHEC - Vérification requise")
    else:
        lines.append("⚠️ Statut: SUCCÈS PARTIEL")
    
    lines.extend(["=" * 60, ""])
    
    resume_text = '\n'.join(lines)
    perf_logger.info(resume_text)
    logger.info(resume_text)


def log_erreur(category, message, exception=None, request=None, **kwargs):
    """Log une erreur avec stack trace"""
    import traceback
    
    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    user_info = "anonymous"
    
    if request and hasattr(request, 'user') and request.user.is_authenticated:
        user_info = request.user.username
    
    extra_info = ' '.join([f"[{k}:{v}]" for k, v in kwargs.items() if v is not None])
    log_msg = f"[{timestamp}] [ERREUR] [{category}] [User:{user_info}] {extra_info} {message}"
    
    if exception:
        log_msg += f"\n  Exception: {type(exception).__name__}: {str(exception)}"
        log_msg += f"\n  Stack trace:\n{traceback.format_exc()}"
    
    logger.error(log_msg)
    anomaly_logger.error(log_msg)


# ================================================================
# FONCTIONS UTILITAIRES JOURS FÉRIÉS
# ================================================================

def get_feries_dans_periode(date_debut, date_fin, code_pays='CI'):
    """
    Retourne la liste des jours fériés dans une période donnée
    """
    from mainapp.models import JourFerie
    
    try:
        feries = JourFerie.objects.filter(
            date_ferie__gte=date_debut,
            date_ferie__lte=date_fin,
            code_pays=code_pays,
            statut='ACTIF'
        ).order_by('date_ferie')
        
        return [
            {
                'date': f.date_ferie,
                'nom': f.nom,
                'type': f.get_type_ferie_display(),
            }
            for f in feries
        ]
        
    except Exception as e:
        log_erreur('FERIES', f"Erreur récupération jours fériés période {date_debut} - {date_fin}", 
                  exception=e, code_pays=code_pays)
        return []


def compter_feries_dans_periode(date_debut, date_fin, code_pays='CI'):
    """Compte le nombre de jours fériés dans une période"""
    from mainapp.models import JourFerie
    
    try:
        return JourFerie.objects.filter(
            date_ferie__gte=date_debut,
            date_ferie__lte=date_fin,
            code_pays=code_pays,
            statut='ACTIF'
        ).count()
    except Exception as e:
        log_erreur('FERIES', f"Erreur comptage jours fériés", exception=e)
        return 0


def formater_feries_pour_affichage(feries):
    """Formate la liste des jours fériés pour l'affichage"""
    if not feries:
        return "-"
    
    return ", ".join([
        f"{f['nom']} ({f['date'].strftime('%d/%m')})"
        for f in feries
    ])


# ================================================================
# VUE PRINCIPALE - LISTE DES ABSENCES
# ================================================================

@login_required
def absences_extraction_liste(request):
    """
    Vue principale pour l'extraction des absences avec jours fériés
    
    URL: /interim/extraction/
    Template: extraction_absences_liste.html
    """
    start_time = time.time()
    
    from mainapp.models import AbsenceUtilisateur, Departement, Site, ProfilUtilisateur
    
    log_action('EXTRACTION', 'ACCES_LISTE', "Accès liste extraction absences", request=request)
    
    # Paramètres de filtrage
    departement_id = request.GET.get('departement', '')
    site_id = request.GET.get('site', '')
    matricule = request.GET.get('matricule', '').strip()
    type_absence = request.GET.get('type_absence', '')
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str = request.GET.get('date_fin', '')
    statut_filtre = request.GET.get('statut', '')
    annee_filtre = request.GET.get('annee', '')
    tri = request.GET.get('tri', 'date_desc')
    
    # Log des filtres actifs
    filtres_actifs = []
    if departement_id:
        filtres_actifs.append(f"dept:{departement_id}")
    if site_id:
        filtres_actifs.append(f"site:{site_id}")
    if matricule:
        filtres_actifs.append(f"matricule:{matricule}")
    if type_absence:
        filtres_actifs.append(f"type:{type_absence}")
    
    if filtres_actifs:
        log_action('EXTRACTION', 'FILTRES_APPLIQUES', 
                  f"Filtres: {', '.join(filtres_actifs)}", request=request)
    
    # Options de tri disponibles
    OPTIONS_TRI = {
        'matricule_asc': ('utilisateur__matricule', 'Matricule (A → Z)'),
        'matricule_desc': ('-utilisateur__matricule', 'Matricule (Z → A)'),
        'nom_asc': ('utilisateur__user__last_name', 'utilisateur__user__first_name', 'Nom (A → Z)'),
        'nom_desc': ('-utilisateur__user__last_name', '-utilisateur__user__first_name', 'Nom (Z → A)'),
        'date_asc': ('date_debut', 'Date début (ancien → récent)'),
        'date_desc': ('-date_debut', 'Date début (récent → ancien)'),
        'motif_asc': ('type_absence', 'Motif (A → Z)'),
        'motif_desc': ('-type_absence', 'Motif (Z → A)'),
    }
    
    today = date.today()
    annee_courante = today.year
    
    # Année sélectionnée
    if annee_filtre:
        try:
            annee_selectionnee = int(annee_filtre)
            if annee_selectionnee < 2000 or annee_selectionnee > 2100:
                log_anomalie('EXTRACTION', f"Année invalide saisie: {annee_filtre}", 
                            severite='INFO', request=request)
                annee_selectionnee = annee_courante
        except ValueError:
            log_anomalie('EXTRACTION', f"Format année invalide: {annee_filtre}", 
                        severite='INFO', request=request)
            annee_selectionnee = annee_courante
    else:
        annee_selectionnee = annee_courante
    
    # Dates
    if not date_debut_str:
        date_debut = date(annee_selectionnee, 1, 1)
    else:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        except ValueError:
            log_anomalie('EXTRACTION', f"Format date début invalide: {date_debut_str}", 
                        severite='INFO', request=request)
            date_debut = date(annee_selectionnee, 1, 1)
    
    if not date_fin_str:
        date_fin = date(annee_selectionnee, 12, 31)
    else:
        try:
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            log_anomalie('EXTRACTION', f"Format date fin invalide: {date_fin_str}", 
                        severite='INFO', request=request)
            date_fin = date(annee_selectionnee, 12, 31)
    
    # Vérification cohérence dates
    if date_debut > date_fin:
        log_anomalie('EXTRACTION', f"Date début ({date_debut}) > Date fin ({date_fin})", 
                    severite='WARNING', request=request)
        date_debut, date_fin = date_fin, date_debut
    
    try:
        # Requête de base
        absences = AbsenceUtilisateur.objects.select_related(
            'utilisateur',
            'utilisateur__departement',
            'utilisateur__site',
            'utilisateur__user'
        ).filter(
            Q(date_debut__lte=date_fin) & Q(date_fin__gte=date_debut)
        )
        
        # Appliquer l'ordre de tri
        if tri in OPTIONS_TRI:
            tri_info = OPTIONS_TRI[tri]
            if tri.startswith('nom_'):
                absences = absences.order_by(tri_info[0], tri_info[1])
            else:
                absences = absences.order_by(tri_info[0])
        else:
            absences = absences.order_by('-date_debut', 'utilisateur__user__last_name')
        
        # Filtrage par département
        if departement_id:
            try:
                absences = absences.filter(utilisateur__departement_id=int(departement_id))
            except ValueError:
                log_anomalie('EXTRACTION', f"ID département invalide: {departement_id}", 
                            severite='WARNING', request=request)
        
        # Filtrage par site
        if site_id:
            try:
                absences = absences.filter(utilisateur__site_id=int(site_id))
            except ValueError:
                log_anomalie('EXTRACTION', f"ID site invalide: {site_id}", 
                            severite='WARNING', request=request)
        
        # Filtrage par matricule
        if matricule:
            absences = absences.filter(
                Q(utilisateur__matricule__icontains=matricule) |
                Q(utilisateur__user__first_name__icontains=matricule) |
                Q(utilisateur__user__last_name__icontains=matricule)
            )
        
        # Filtrage par type d'absence
        if type_absence:
            absences = absences.filter(type_absence__icontains=type_absence)
        
        # Enrichir avec les jours fériés
        absences_avec_feries = []
        total_feries = 0
        for absence in absences:
            feries = get_feries_dans_periode(absence.date_debut, absence.date_fin)
            nb_feries = len(feries)
            total_feries += nb_feries
            absences_avec_feries.append({
                'absence': absence,
                'feries': feries,
                'nb_feries': nb_feries,
                'feries_texte': formater_feries_pour_affichage(feries),
            })
        
    except Exception as e:
        log_erreur('EXTRACTION', "Erreur requête absences", exception=e, request=request)
        absences_avec_feries = []
        total_feries = 0
    
    # Pagination
    page_number = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', '25')
    
    OPTIONS_PER_PAGE = [10, 25, 50, 100, 200]
    
    try:
        per_page = int(per_page)
        if per_page not in OPTIONS_PER_PAGE:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25
    
    paginator = Paginator(absences_avec_feries, per_page)
    page_obj = paginator.get_page(page_number)
    
    # Données pour les filtres
    try:
        departements = Departement.objects.filter(actif=True).order_by('nom')
        sites = Site.objects.filter(actif=True).order_by('nom')
        types_absence = AbsenceUtilisateur.objects.values_list(
            'type_absence', flat=True
        ).distinct().order_by('type_absence')
    except Exception as e:
        log_erreur('EXTRACTION', "Erreur chargement filtres", exception=e, request=request)
        departements = []
        sites = []
        types_absence = []
    
    # Statistiques
    stats = {
        'total_absences': len(absences_avec_feries),
        'total_jours': sum(a['absence'].duree_jours for a in absences_avec_feries),
        'absences_avec_feries': sum(1 for a in absences_avec_feries if a['nb_feries'] > 0),
        'total_feries_inclus': total_feries,
    }
    
    # Options de tri pour le template
    options_tri_template = [
        ('date_desc', 'Date début (récent → ancien)'),
        ('date_asc', 'Date début (ancien → récent)'),
        ('matricule_asc', 'Matricule (A → Z)'),
        ('matricule_desc', 'Matricule (Z → A)'),
        ('nom_asc', 'Nom (A → Z)'),
        ('nom_desc', 'Nom (Z → A)'),
        ('motif_asc', 'Motif (A → Z)'),
        ('motif_desc', 'Motif (Z → A)'),
    ]
    
    # Log résumé
    duree_ms = (time.time() - start_time) * 1000
    
    log_resume('EXTRACTION_ABSENCES_LISTE', {
        'utilisateur': request.user.username,
        'periode': f"{date_debut} à {date_fin}",
        'total_absences': stats['total_absences'],
        'total_jours': stats['total_jours'],
        'absences_avec_feries': stats['absences_avec_feries'],
        'filtres_actifs': len(filtres_actifs),
        'page': page_number,
    }, duree_ms=duree_ms)
    
    # Détection anomalies
    if stats['total_absences'] == 0 and not filtres_actifs:
        log_anomalie('EXTRACTION', f"Aucune absence trouvée pour l'année {annee_selectionnee}", 
                    severite='INFO', request=request)
    
    if duree_ms > 5000:
        log_anomalie('PERFORMANCE', f"Chargement lent extraction: {duree_ms:.0f}ms", 
                    severite='WARNING', request=request, total_absences=stats['total_absences'])
    
    context = {
        'page_obj': page_obj,
        'absences': page_obj.object_list,
        'departements': departements,
        'sites': sites,
        'types_absence': types_absence,
        'stats': stats,
        'departement_id': departement_id,
        'site_id': site_id,
        'matricule': matricule,
        'type_absence': type_absence,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_filtre': statut_filtre,
        'annee_selectionnee': annee_selectionnee,
        'annee_courante': annee_courante,
        'tri': tri,
        'options_tri': options_tri_template,
        'per_page': per_page,
        'options_per_page': OPTIONS_PER_PAGE,
        'today': today,
    }
    
    return render(request, 'extraction_absences_liste.html', context)


# ================================================================
# VUE EXPORT - PDF, XLSX, CSV
# ================================================================

@login_required
def absences_extraction_export(request, format_export):
    """
    Export des absences en PDF, XLSX ou CSV
    
    URL: /interim/extraction/export/<format>/
    Formats: pdf, xlsx, csv
    """
    start_time = time.time()
    
    from mainapp.models import AbsenceUtilisateur
    
    log_action('EXPORT', 'DEBUT_EXPORT', f"Début export {format_export.upper()}", 
              request=request, format=format_export)
    
    # Valider le format
    if format_export not in ['pdf', 'xlsx', 'csv']:
        log_anomalie('EXPORT', f"Format d'export non supporté: {format_export}", 
                    severite='WARNING', request=request)
        return HttpResponse("Format non supporté", status=400)
    
    # Récupérer les filtres
    departement_id = request.GET.get('departement', '')
    site_id = request.GET.get('site', '')
    matricule = request.GET.get('matricule', '').strip()
    type_absence = request.GET.get('type_absence', '')
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str = request.GET.get('date_fin', '')
    annee_filtre = request.GET.get('annee', '')
    tri = request.GET.get('tri', 'date_desc')
    
    OPTIONS_TRI = {
        'matricule_asc': ('utilisateur__matricule',),
        'matricule_desc': ('-utilisateur__matricule',),
        'nom_asc': ('utilisateur__user__last_name', 'utilisateur__user__first_name'),
        'nom_desc': ('-utilisateur__user__last_name', '-utilisateur__user__first_name'),
        'date_asc': ('date_debut',),
        'date_desc': ('-date_debut',),
        'motif_asc': ('type_absence',),
        'motif_desc': ('-type_absence',),
    }
    
    today = date.today()
    annee_courante = today.year
    
    try:
        annee_selectionnee = int(annee_filtre) if annee_filtre else annee_courante
        if annee_selectionnee < 2000 or annee_selectionnee > 2100:
            annee_selectionnee = annee_courante
    except ValueError:
        annee_selectionnee = annee_courante
    
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date() if date_debut_str else date(annee_selectionnee, 1, 1)
    except ValueError:
        date_debut = date(annee_selectionnee, 1, 1)
    
    try:
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date() if date_fin_str else date(annee_selectionnee, 12, 31)
    except ValueError:
        date_fin = date(annee_selectionnee, 12, 31)
    
    try:
        absences = AbsenceUtilisateur.objects.select_related(
            'utilisateur',
            'utilisateur__departement',
            'utilisateur__site',
            'utilisateur__user'
        ).filter(
            Q(date_debut__lte=date_fin) & Q(date_fin__gte=date_debut)
        )
        
        if tri in OPTIONS_TRI:
            absences = absences.order_by(*OPTIONS_TRI[tri])
        else:
            absences = absences.order_by('-date_debut', 'utilisateur__user__last_name')
        
        if departement_id:
            try:
                absences = absences.filter(utilisateur__departement_id=int(departement_id))
            except ValueError:
                pass
        
        if site_id:
            try:
                absences = absences.filter(utilisateur__site_id=int(site_id))
            except ValueError:
                pass
        
        if matricule:
            absences = absences.filter(
                Q(utilisateur__matricule__icontains=matricule) |
                Q(utilisateur__user__first_name__icontains=matricule) |
                Q(utilisateur__user__last_name__icontains=matricule)
            )
        
        if type_absence:
            absences = absences.filter(type_absence__icontains=type_absence)
        
        total_absences = absences.count()
        
        if total_absences > 10000:
            log_anomalie('EXPORT', f"Export volumineux: {total_absences} enregistrements", 
                        severite='WARNING', request=request, format=format_export)
        
        if total_absences == 0:
            log_anomalie('EXPORT', "Export sans données - aucune absence trouvée", 
                        severite='INFO', request=request)
        
        donnees = []
        total_avec_feries = 0
        
        for absence in absences:
            feries = get_feries_dans_periode(absence.date_debut, absence.date_fin)
            nb_feries = len(feries)
            if nb_feries > 0:
                total_avec_feries += 1
            
            donnees.append({
                'matricule': absence.utilisateur.matricule,
                'nom': absence.utilisateur.user.last_name if absence.utilisateur.user else '',
                'prenom': absence.utilisateur.user.first_name if absence.utilisateur.user else '',
                'departement': absence.utilisateur.departement.nom if absence.utilisateur.departement else '',
                'site': absence.utilisateur.site.nom if absence.utilisateur.site else '',
                'date_debut': absence.date_debut,
                'date_fin': absence.date_fin,
                'duree_jours': absence.duree_jours,
                'type_absence': absence.type_absence,
                'feries': formater_feries_pour_affichage(feries),
                'nb_feries': nb_feries,
            })
        
        filename = f"extraction_absences_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}"
        
        if format_export == 'csv':
            response = export_csv(donnees, filename)
        elif format_export == 'xlsx':
            response = export_xlsx(donnees, filename, date_debut, date_fin)
        elif format_export == 'pdf':
            response = export_pdf(donnees, filename, date_debut, date_fin)
        else:
            response = HttpResponse("Format non supporté", status=400)
        
        duree_ms = (time.time() - start_time) * 1000
        
        log_action('EXPORT', 'EXPORT_TERMINE', 
                  f"Export {format_export.upper()} terminé: {len(donnees)} lignes",
                  request=request, format=format_export, lignes=len(donnees))
        
        log_resume(f'EXPORT_ABSENCES_{format_export.upper()}', {
            'utilisateur': request.user.username,
            'format': format_export.upper(),
            'periode': f"{date_debut} à {date_fin}",
            'total_absences': len(donnees),
            'absences_avec_feries': total_avec_feries,
            'fichier': f"{filename}.{format_export}",
            'statut': 'SUCCÈS'
        }, duree_ms=duree_ms)
        
        return response
        
    except Exception as e:
        duree_ms = (time.time() - start_time) * 1000
        log_erreur('EXPORT', f"Erreur export {format_export}", exception=e, request=request)
        
        log_resume(f'EXPORT_ABSENCES_{format_export.upper()}', {
            'utilisateur': request.user.username,
            'format': format_export.upper(),
            'statut': 'ÉCHEC',
            'erreur': str(e)[:100]
        }, duree_ms=duree_ms)
        
        return HttpResponse(f"Erreur lors de l'export: {str(e)}", status=500)


# ================================================================
# FONCTIONS D'EXPORT
# ================================================================

def export_csv(donnees, filename):
    """Export en CSV"""
    log_action('EXPORT', 'GENERATION_CSV', f"Génération CSV: {len(donnees)} lignes")
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'MATRICULE', 'NOM', 'PRENOM', 'DEPARTEMENT', 'SITE',
        'DATE DEPART', 'DATE FIN', 'NOMBRE DE JOURS',
        'MOTIF DE L\'ABSENCE', 'JOURS FERIES INCLUS', 'NB FERIES'
    ])
    
    for d in donnees:
        writer.writerow([
            d['matricule'], d['nom'], d['prenom'], d['departement'], d['site'],
            d['date_debut'].strftime('%d/%m/%Y'), d['date_fin'].strftime('%d/%m/%Y'),
            d['duree_jours'], d['type_absence'], d['feries'], d['nb_feries']
        ])
    
    return response


def export_xlsx(donnees, filename, date_debut, date_fin):
    """Export en XLSX"""
    log_action('EXPORT', 'GENERATION_XLSX', f"Génération XLSX: {len(donnees)} lignes")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_anomalie('EXPORT', "Module openpyxl non installé", severite='ERROR')
        return HttpResponse("Module openpyxl non installé", status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Absences"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ferie_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws.merge_cells('A1:K1')
    ws['A1'] = f"EXTRACTION DES ABSENCES - DU {date_debut.strftime('%d/%m/%Y')} AU {date_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:K2')
    ws['A2'] = f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True, size=10)
    
    headers = ['MATRICULE', 'NOM', 'PRENOM', 'DEPARTEMENT', 'SITE', 'DATE DEPART',
               'DATE FIN', 'NB JOURS', 'MOTIF', 'JOURS FERIES', 'NB FERIES']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, d in enumerate(donnees, 5):
        row_data = [
            d['matricule'], d['nom'], d['prenom'], d['departement'], d['site'],
            d['date_debut'].strftime('%d/%m/%Y'), d['date_fin'].strftime('%d/%m/%Y'),
            d['duree_jours'], d['type_absence'], d['feries'], d['nb_feries']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if d['nb_feries'] > 0 and col in [10, 11]:
                cell.fill = ferie_fill
    
    column_widths = [12, 15, 15, 20, 15, 12, 12, 10, 25, 35, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = 'A5'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def export_pdf(donnees, filename, date_debut, date_fin):
    """Export en PDF"""
    log_action('EXPORT', 'GENERATION_PDF', f"Génération PDF: {len(donnees)} lignes")
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        log_anomalie('EXPORT', "Module reportlab non installé", severite='ERROR')
        return HttpResponse("Module reportlab non installé", status=500)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                 fontSize=14, alignment=TA_CENTER, spaceAfter=20)
    
    title = f"EXTRACTION DES ABSENCES<br/>Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(title, title_style))
    
    sub_style = ParagraphStyle('SubTitle', parent=styles['Normal'],
                               fontSize=9, alignment=TA_CENTER, spaceAfter=15)
    elements.append(Paragraph(f"Extrait le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_style))
    
    table_data = [['Matricule', 'Nom', 'Prénom', 'Département', 'Date Départ',
                   'Date Fin', 'Jours', 'Motif', 'Fériés']]
    
    for d in donnees:
        feries_court = d['feries'] if len(d['feries']) < 30 else d['feries'][:30] + '...'
        table_data.append([
            d['matricule'], d['nom'], d['prenom'],
            d['departement'][:15] if d['departement'] else '',
            d['date_debut'].strftime('%d/%m/%Y'), d['date_fin'].strftime('%d/%m/%Y'),
            str(d['duree_jours']),
            d['type_absence'][:20] if d['type_absence'] else '',
            feries_court if d['nb_feries'] > 0 else '-'
        ])
    
    col_widths = [2*cm, 2.5*cm, 2.5*cm, 3*cm, 2.5*cm, 2.5*cm, 1.5*cm, 4*cm, 5*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (4, 1), (6, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ])
    
    for row_idx, d in enumerate(donnees, 1):
        if d['nb_feries'] > 0:
            table_style.add('BACKGROUND', (8, row_idx), (8, row_idx), colors.HexColor('#FFF2CC'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    elements.append(Spacer(1, 20))
    stats_text = f"Total: {len(donnees)} absences | Absences avec fériés: {sum(1 for d in donnees if d['nb_feries'] > 0)}"
    elements.append(Paragraph(stats_text, styles['Normal']))
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


# ================================================================
# API JSON
# ================================================================

@login_required
def api_absences_extraction(request):
    """
    API JSON pour les absences avec jours fériés
    
    URL: /interim/api/extraction/
    """
    start_time = time.time()
    
    from mainapp.models import AbsenceUtilisateur
    
    log_action('API', 'ACCES_API_EXTRACTION', "Accès API extraction absences", request=request)
    
    departement_id = request.GET.get('departement', '')
    site_id = request.GET.get('site', '')
    matricule = request.GET.get('matricule', '').strip()
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str = request.GET.get('date_fin', '')
    annee_filtre = request.GET.get('annee', '')
    
    today = date.today()
    annee_courante = today.year
    
    try:
        annee_selectionnee = int(annee_filtre) if annee_filtre else annee_courante
    except ValueError:
        annee_selectionnee = annee_courante
    
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date() if date_debut_str else date(annee_selectionnee, 1, 1)
    except ValueError:
        date_debut = date(annee_selectionnee, 1, 1)
    
    try:
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date() if date_fin_str else date(annee_selectionnee, 12, 31)
    except ValueError:
        date_fin = date(annee_selectionnee, 12, 31)
    
    try:
        absences = AbsenceUtilisateur.objects.select_related(
            'utilisateur', 'utilisateur__departement', 'utilisateur__user'
        ).filter(
            Q(date_debut__lte=date_fin) & Q(date_fin__gte=date_debut)
        ).order_by('-date_debut')
        
        if departement_id:
            try:
                absences = absences.filter(utilisateur__departement_id=int(departement_id))
            except ValueError:
                pass
        
        if site_id:
            try:
                absences = absences.filter(utilisateur__site_id=int(site_id))
            except ValueError:
                pass
        
        if matricule:
            absences = absences.filter(
                Q(utilisateur__matricule__icontains=matricule) |
                Q(utilisateur__user__first_name__icontains=matricule) |
                Q(utilisateur__user__last_name__icontains=matricule)
            )
        
        total_count = absences.count()
        limite = 100
        
        if total_count > limite:
            log_anomalie('API', f"Résultats API limités: {total_count} → {limite}", 
                        severite='INFO', request=request)
        
        data = []
        for absence in absences[:limite]:
            feries = get_feries_dans_periode(absence.date_debut, absence.date_fin)
            data.append({
                'id': absence.pk,
                'matricule': absence.utilisateur.matricule,
                'nom_complet': absence.utilisateur.nom_complet,
                'departement': absence.utilisateur.departement.nom if absence.utilisateur.departement else None,
                'date_debut': absence.date_debut.isoformat(),
                'date_fin': absence.date_fin.isoformat(),
                'duree_jours': absence.duree_jours,
                'type_absence': absence.type_absence,
                'feries': [{'date': f['date'].isoformat(), 'nom': f['nom']} for f in feries],
                'nb_feries': len(feries),
            })
        
        duree_ms = (time.time() - start_time) * 1000
        
        log_resume('API_EXTRACTION_ABSENCES', {
            'utilisateur': request.user.username,
            'periode': f"{date_debut} à {date_fin}",
            'total_resultats': total_count,
            'resultats_retournes': len(data),
        }, duree_ms=duree_ms)
        
        return JsonResponse({
            'absences': data,
            'count': len(data),
            'total_count': total_count,
            'date_debut': date_debut.isoformat(),
            'date_fin': date_fin.isoformat(),
        })
        
    except Exception as e:
        log_erreur('API', "Erreur API extraction absences", exception=e, request=request)
        return JsonResponse({
            'error': str(e),
            'absences': [],
            'count': 0
        }, status=500)